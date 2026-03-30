import pandas as pd
import openpyxl
from datetime import datetime
import uuid
from pathlib import Path

WORKBOOK_FILENAME = "Pengeluaran_budget_template.xlsx"
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK_PATH = PROJECT_ROOT / WORKBOOK_FILENAME

TRANSACTION_HEADERS = [
    "transaction_id",
    "date",
    "month",
    "merchant",
    "description",
    "amount",
    "type",
    "category",
    "subcategory",
    "source",
    "receipt_file",
    "receipt_text",
    "notes",
    "created_at",
]


def get_default_workbook_path():
    """Return the committed workbook path used by the app."""
    return str(DEFAULT_WORKBOOK_PATH)


def ensure_transactions_sheet(workbook):
    """Ensure the workbook has a Transactions sheet with the expected headers."""
    if "Transactions" not in workbook.sheetnames:
        sheet = workbook.create_sheet("Transactions")
        for col, header in enumerate(TRANSACTION_HEADERS, 1):
            sheet.cell(row=1, column=col, value=header)
        return sheet

    sheet = workbook["Transactions"]
    if sheet.max_row == 0:
        for col, header in enumerate(TRANSACTION_HEADERS, 1):
            sheet.cell(row=1, column=col, value=header)
        return sheet

    existing_headers = [sheet.cell(row=1, column=col).value for col in range(1, len(TRANSACTION_HEADERS) + 1)]
    if existing_headers != TRANSACTION_HEADERS:
        for col, header in enumerate(TRANSACTION_HEADERS, 1):
            sheet.cell(row=1, column=col, value=header)
    return sheet

def load_workbook(file_path):
    """Load the Excel workbook."""
    workbook = openpyxl.load_workbook(file_path)
    # Store source path for compatibility with openpyxl versions
    # where `Workbook.filename` may not be available.
    workbook._file_path = file_path
    ensure_transactions_sheet(workbook)
    return workbook

def get_transactions_df(workbook):
    """Get the Transactions sheet as a DataFrame."""
    sheet = ensure_transactions_sheet(workbook)
    data = list(sheet.values)
    if len(data) <= 1:
        return pd.DataFrame(columns=data[0] if data else TRANSACTION_HEADERS)
    df = pd.DataFrame(data[1:], columns=data[0])
    if "amount" in df.columns:
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
    if "month" in df.columns:
        df["month"] = df["month"].fillna("")
    return df

def get_budget_df(workbook):
    """Get the Budget Template sheet as a DataFrame."""
    if "Budget Template" not in workbook.sheetnames:
        raise ValueError("Budget Template sheet not found")
    sheet = workbook["Budget Template"]
    data = list(sheet.values)
    df = pd.DataFrame(data[1:], columns=data[0])
    return df

def append_transaction(workbook, transaction_data):
    """Append a transaction to the Transactions sheet."""
    sheet = ensure_transactions_sheet(workbook)
    # Generate transaction_id if not provided
    if "transaction_id" not in transaction_data:
        transaction_data["transaction_id"] = str(uuid.uuid4())
    if "created_at" not in transaction_data:
        transaction_data["created_at"] = datetime.now()
    if "month" not in transaction_data and "date" in transaction_data:
        transaction_data["month"] = transaction_data["date"].strftime("%Y-%m") if isinstance(transaction_data["date"], datetime) else str(transaction_data["date"])[:7]

    row = [transaction_data.get(header, "") for header in TRANSACTION_HEADERS]
    sheet.append(row)
    file_path = getattr(workbook, "_file_path", None)
    if file_path:
        workbook.save(file_path)

def replace_transactions(workbook, transactions_df):
    """Replace the Transactions sheet contents and save the workbook."""
    sheet = ensure_transactions_sheet(workbook)
    sheet.delete_rows(2, sheet.max_row)
    normalized_df = transactions_df.reindex(columns=TRANSACTION_HEADERS, fill_value="")
    for row in normalized_df.itertuples(index=False, name=None):
        sheet.append(list(row))
    save_workbook(workbook)


def save_workbook(workbook, file_path=None):
    """Save the workbook."""
    if file_path is None:
        file_path = getattr(workbook, "_file_path", None)
    if file_path is None:
        raise ValueError("Workbook path is required to save the workbook")
    workbook.save(file_path)
